# -*- codeing = utf-8 -*-
from bs4 import BeautifulSoup  # 网页解析，获取数据
import re  # 正则表达式，进行文字匹配`
import urllib.request, urllib.error  # 制定URL，获取网页数据
import xlwt  # 进行excel操作
import urllib.parse
import time
import openpyxl

nwb=openpyxl.Workbook()
ws=nwb.active

def main():
    getData()

# 爬取网页
def getData():
    findwords = re.compile('[\u4e00-\u9fa5]{1,}')
    findpinyin = re.compile(r'<rt>(.*?)<')  # 创建查找 拼音的正则表达式对象
    findshiyi = re.compile(r'<span class="gc_sy">(.*?)</span>')  # 创建查找 释义的正则表达式对象
    myfile = open('词汇.txt', encoding='utf-8')
    str1 = myfile.read()
    datalist = []  #用来存储爬取的网页信息
    words = findwords.findall(str1)
    x = 1
    for word in words:
        time.sleep(1)
        baseurl = "https://www.zdic.net/hans/" + urllib.parse.quote(word)  # 要爬取的网页链接,用urllib.parse.quote(word)解决中文路径问题，需要import urllib.parse
        html = askURL(baseurl)  # 保存获取到的网页源码
        # 2.逐一解析数据
        soup = BeautifulSoup(html, "html.parser")
        for item in soup.find_all('div', class_="gnr"):  # 查找符合要求的字符串
            item = str(item)
            ws.cell(x, 1, word)
            pinyin = re.findall(findpinyin, item)[0]  # 查找拼音
            ws.cell(x, 2, pinyin)
            shiyi = re.findall(findshiyi,item)[0]  # 查找释义
            ws.cell(x, 3, shiyi)
            x += 1
    nwb.save('cihui.xlsx')  # 保存工作簿。

# 得到指定一个URL的网页内容
def askURL(url):
    head = {  # 模拟浏览器头部信息，向服务器发送消息
        "User-Agent": "Mozilla / 5.0(Windows NT 10.0; Win64; x64) AppleWebKit / 537.36(KHTML, like Gecko) Chrome / 80.0.3987.122  Safari / 537.36"
    }
    # 用户代理，表示告诉服务器，我们是什么类型的机器、浏览器（本质上是告诉浏览器，我们可以接收什么水平的文件内容）

    request = urllib.request.Request(url, headers=head)
    html = ""
    try:
        response = urllib.request.urlopen(request)
        html = response.read().decode("utf-8")
    except urllib.error.URLError as e:
        if hasattr(e, "code"):
            print(e.code)
        if hasattr(e, "reason"):
            print(e.reason)
    return html

if __name__ == "__main__":  # 当程序执行时
    # 调用函数
     main()
    # init_db("movietest.db")
     print("爬取完毕！")