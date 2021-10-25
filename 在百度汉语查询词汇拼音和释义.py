#百度汉语查词汇
# -*- codeing = utf-8 -*-
from bs4 import BeautifulSoup  # 网页解析，获取数据
import re  # 正则表达式，进行文字匹配`
import urllib.request, urllib.error  # 制定URL，获取网页数据
import xlwt  # 进行excel操作
import urllib.parse
import time
import openpyxl

def main():
    getData()

# 爬取网页
def getData():

    findwords = re.compile('[\u4e00-\u9fa5]{1,}') #提取汉字的正则表达式
    findpinyin = re.compile(r'<dt class="pinyin">\[ (.*?) \]</dt>', re.DOTALL)  # 创建查找 拼音的正则表达式对象
    findshiyi = re.compile(r'<p>(.*?)</p>',re.DOTALL)  # 创建查找 释义的正则表达式对象

    wb = openpyxl.load_workbook('cihui.xlsx')  # 读取工作簿。
    #ws = wb.worksheets[0]  # 读取工作表。
    ws = wb['cihui']  # 读取工作表。
    now_row=2 #excel当前行
    #kk=list(ws.values)
    while now_row <= ws.max_row:   #for row in list(ws.values)[1:]: #从第二行逐行读取表数据，这种方式似乎更好
        word=ws.cell(now_row,2).value
        print("正在查询：{}".format(word))
        #time.sleep(0.2)
        baseurl = "https://hanyu.baidu.com/zici/s?wd=" + urllib.parse.quote(word) + "&from=zici"  # 要爬取的网页链接,用urllib.parse.quote(word)解决中文路径问题，需要import urllib.parse
        html = askURL(baseurl)  # 保存获取到的网页源码
        # 2.逐一解析数据
        soup = BeautifulSoup(html, "html.parser")
        # for item in soup.find_all('div', class_="poem-detail-body"):  # 查找符合要求的字符串
        for item in soup.find_all('div', id="basicmean-wrapper"):  # 查找符合要求的字符串
            item = str(item)
            try:
                pinyin = re.findall(findpinyin, item)[0]  # 查找拼音
            except:
                print("抓取 {} 拼音失败".format(word))
            else:
                ws.cell(now_row, 3, pinyin)
            try:
                shiyi=''  #有些释义含多行，需要循环提取成字符串
                for shy in re.findall(findshiyi,item):  # 查找释义
                    shiyi=shiyi+shy
            except:
                print("查询 {} 释义失败".format(word))
            else:
                tr = re.compile('[^\n\s]') #删除空格和回车，返回值是列表
                ks=re.findall(tr,shiyi)
                shiyi1=''
                for i in ks:
                    shiyi1=shiyi1+i   #将列表制作成字符串
                ws.cell(now_row, 4, shiyi1)
        now_row += 1
    wb.save('cihui.xlsx')  # 保存工作簿。

# 得到指定一个URL的网页内容
def askURL(url):
    head = {  # 模拟浏览器头部信息，向服务器发送消息
        "User-Agent": "Mozilla / 5.0(Windows NT 10.0; Win64; x64) AppleWebKit / 537.36(KHTML, like Gecko) Chrome / 80.0.3987.122  Safari / 537.36"
    }
    # 用户代理，表示告诉服务器，我们是什么类型的机器、浏览器（本质上是告诉浏览器，我们可以接收什么水平的文件内容）

    request = urllib.request.Request(url, headers=head)
    html = ""
    try:
        response = urllib.request.urlopen(request)
        html = response.read().decode("utf-8","ignore" )  #在查询"薄纸，一场雨"等词汇出现程序中断，出现异常报错是由于设置了decode()方法的第二个参数errors为严格（strict）形式造成的，因为默认就是这个参数，将其更改为ignore等即可。
    except urllib.error.URLError as e:
        if hasattr(e, "code"):
            print(e.code)
        if hasattr(e, "reason"):
            print(e.reason)
    return html

if __name__ == "__main__":  # 当程序执行时
    # 调用函数
     main()
    # init_db("movietest.db")
     print("查询结束！")